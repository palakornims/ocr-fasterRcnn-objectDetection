from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import abc
import openpyxl
from openpyxl import Workbook
import sys as system
import collections
from object_detection.core import standard_fields as fields
from object_detection.utils import shape_utils

def string_sort(
    boxes,
    classes,
    scores,
    category_index,
    max_boxes_to_draw=20,
    min_score_thresh=.5,):


  list_xmin = []
  class_name=[]

  if not max_boxes_to_draw:
    max_boxes_to_draw = boxes.shape[0]
  for i in range(min(max_boxes_to_draw, boxes.shape[0])):

    if scores is None or scores[i] > min_score_thresh:

      box = tuple(boxes[i].tolist())
      ymin, xmin, ymax, xmax = box
      class_name.append(category_index[classes[i]]['name'])
      list_xmin.append(xmin)
      lists=[]
      lists= sort_lst(list_xmin)
      string_sort = []
      string_sort=string_sort_list(class_name,list_xmin,lists)
      class_name2 = category_index[classes[i]]['name']
      

  return string_sort


def sort_lst(list):
  sort_list=list.copy()

  sort_list.sort()

  return sort_list

def string_sort_list(string_list,xmin_list,xmin_sort_list):
  string_sort_list=[]

  for i in xmin_sort_list:
    for j in xmin_list:
      if i == j:
        position=xmin_list.index(i)
        string_sort_list.append(string_list[position])



  return string_sort_list


def check_sum(ocr):
  checksumref=0
  print(ocr)
  ocrcondition=False
  ocrbytes=str.encode(ocr)
  list_byte_str=list(ocrbytes)
  print(list_byte_str)
  for i in list_byte_str:
    
    checksumref=((checksumref*8)+(i-32))%59
    
  if checksumref==0:
    ocrcondition=True
  else:
    ocrcondition=ocrcondition

  return ocrcondition


def convert_list_to_str(strlist):
  ocrstring = ''
  for cat in strlist:
    ocrstring = ocrstring+cat

  return ocrstring



def write_to_excel(image_name,ocr_string,ocr_condition):

     wb_row_check=openpyxl.load_workbook('C:/tensorflow2/OCR log/Bad_ocr_rcnn.xlsx')
     row_check_sheet = wb_row_check.active
     last_row =row_check_sheet.max_row






     if row_check_sheet.cell(column=1, row=last_row+1).value is None and last_row > 0:
       last_row=last_row+1



     row_check_sheet['A'+str(last_row)]=image_name
     row_check_sheet['B'+str(last_row)]=ocr_string
     row_check_sheet['C'+str(last_row)]=ocr_condition






     wb_row_check.save('C:/tensorflow2/OCR log/Bad_ocr_rcnn.xlsx')




